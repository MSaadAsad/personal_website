#!/usr/bin/env python3
"""Build compact district demographic and housing indicators from PBS Census 2023 tables."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

import openpyxl


AREAS = ("punjab", "sindh", "kp", "balochistan", "islamabad")
ALIASES = {
    "central karachi": "karachicentral",
    "east karachi": "karachieast",
    "south karachi": "karachisouth",
    "west karachi": "karachiwest",
    "malir": "karachimalir",
    "korangi": "karachikorangi",
    "killa abdullah": "qillaabdullah",
    "killa saifullah": "qillasaifullah",
    "shaheed benazir abad": "shaheedbenazirabad",
    "chagai": "chaghi",
    "lower chitral": "chitrallower",
    "upper chitral": "chitralupper",
    "lower kohistan": "kohistanlower",
    "upper kohistan": "kohistanupper",
    "tando ahyar": "tandoallahyar",
}


def district_key(name: str) -> str:
    clean = re.sub(r"\s+district\s*$", "", name.strip(), flags=re.I)
    alias = ALIASES.get(clean.lower())
    return alias or re.sub(r"[^a-z0-9]", "", clean.lower())


def number(value):
    return value if isinstance(value, (int, float)) else None


def workbook_path(root: Path, table: int, area: str) -> Path:
    suffix = "islamabad" if area == "islamabad" else f"{area}_districts"
    return root / f"table_{table}_{suffix}.xlsx"


def iter_rows(path: Path):
    sheet = openpyxl.load_workbook(path, data_only=True, read_only=True).active
    yield from sheet.iter_rows(values_only=True)


def parse_table_1(path: Path, province: str, result: dict):
    for row in iter_rows(path):
        label = str(row[0] or "").strip()
        if not re.search(r"\bDISTRICT$", label, re.I):
            continue
        key = district_key(label)
        result[key] = {
            "name": re.sub(r"\s+district$", "", label, flags=re.I).title(),
            "province": province,
            "areaKm2": number(row[1]),
            "population": number(row[2]),
            "male": number(row[3]),
            "female": number(row[4]),
            "sexRatio": number(row[6]),
            "density": number(row[7]),
            "householdSize": number(row[9]),
            "growthRate": number(row[11]),
        }


def parse_table_5(path: Path, result: dict):
    current = None
    for row in iter_rows(path):
        label = str(row[0] or "").strip()
        if re.search(r"\bDISTRICT$", label, re.I):
            current = district_key(label)
            continue
        if current not in result:
            continue
        value = number(row[1])
        if value is None:
            continue
        if label == "ALL AGES":
            result[current]["agePopulation"] = value
        elif label == "UNDER 15":
            result[current]["under15"] = value
        elif label == "15 -- 64":
            result[current]["age15to64"] = value
        elif label == "65 &  ABOVE":
            result[current]["age65plus"] = value


def parse_households(path: Path, result: dict, columns: dict[str, int]):
    current = None
    for row in iter_rows(path):
        label = str(row[0] or "").strip()
        if re.search(r"\bDISTRICT$", label, re.I):
            current = district_key(label)
            continue
        if label != "ALL LOCALITIES" or current not in result:
            continue
        for field, column in columns.items():
            result[current][field] = number(row[column])


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, default=Path("/tmp/pbs-census"))
    parser.add_argument("--output", type=Path, default=Path("public/data/pakistan-map/census-2023-detail.json"))
    args = parser.parse_args()
    districts: dict[str, dict] = {}
    province_names = {"punjab": "Punjab", "sindh": "Sindh", "kp": "Khyber Pakhtunkhwa", "balochistan": "Balochistan", "islamabad": "Islamabad"}

    for area in AREAS:
        parse_table_1(workbook_path(args.input, 1, area), province_names[area], districts)
        parse_table_5(workbook_path(args.input, 5, area), districts)
        parse_households(workbook_path(args.input, 23, area), districts, {
            "waterHouseholds": 1, "improvedWater": 2, "waterInside": 3, "tapWater": 5,
        })
        parse_households(workbook_path(args.input, 24, area), districts, {
            "sanitationHouseholds": 1, "flushToilet": 3, "noToilet": 5, "separateWashroom": 6,
        })
        parse_households(workbook_path(args.input, 25, area), districts, {
            "housingHouseholds": 1, "ownedHouse": 2, "rentedHouse": 3, "femaleOwner": 9, "oneRoom": 10,
        })

    payload = {
        "source": "Pakistan Bureau of Statistics, Population and Housing Census 2023",
        "sourceUrl": "https://www.pbs.gov.pk/result-excel/",
        "tables": [1, 5, 23, 24, 25],
        "coverage": "Punjab, Sindh, Khyber Pakhtunkhwa, Balochistan and Islamabad; district rows only",
        "districts": dict(sorted(districts.items())),
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, separators=(",", ":")), encoding="utf-8")
    print(f"Wrote {len(districts)} districts to {args.output}")


if __name__ == "__main__":
    main()
